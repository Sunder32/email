import csv
import io
import re

from openpyxl import load_workbook

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
HEADER_KEYWORDS = {"email", "e-mail", "mail", "почта", "адрес", "address"}


def _extract_email(cell) -> str | None:
    if cell is None:
        return None
    value = str(cell).strip().lower()
    if not value:
        return None
    return value if EMAIL_RE.match(value) else None


def parse_emails_from_csv(content: bytes) -> list[str]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader if row]
    if not rows:
        return []

    email_col = 0
    first_row = rows[0]
    first_is_header = any(
        str(c).strip().lower() in HEADER_KEYWORDS for c in first_row
    )

    if first_is_header:
        for j, cell in enumerate(first_row):
            if str(cell).strip().lower() in HEADER_KEYWORDS:
                email_col = j
                break
        data_rows = rows[1:]
    else:
        for j, cell in enumerate(first_row):
            if _extract_email(cell):
                email_col = j
                break
        data_rows = rows

    emails: list[str] = []
    for row in data_rows:
        if email_col < len(row):
            email = _extract_email(row[email_col])
            if email:
                emails.append(email)
    return emails


def parse_emails_from_xlsx(content: bytes) -> list[str]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    email_col = 0
    first_row = rows[0]
    first_is_header = any(
        str(c or "").strip().lower() in HEADER_KEYWORDS for c in first_row
    )

    if first_is_header:
        for j, cell in enumerate(first_row):
            if str(cell or "").strip().lower() in HEADER_KEYWORDS:
                email_col = j
                break
        data_rows = rows[1:]
    else:
        for j, cell in enumerate(first_row):
            if _extract_email(cell):
                email_col = j
                break
        data_rows = rows

    emails: list[str] = []
    for row in data_rows:
        if email_col < len(row):
            email = _extract_email(row[email_col])
            if email:
                emails.append(email)
    return emails


def parse_emails(filename: str, content: bytes) -> list[str]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        return parse_emails_from_xlsx(content)
    return parse_emails_from_csv(content)
